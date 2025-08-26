"""Unit tests for Excel writer module."""

import pytest
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

from src.auto_scrape.core.config import ExcelConfig
from src.auto_scrape.excel.writer import ExcelWriter


class TestExcelWriter:
    """Test ExcelWriter class."""
    
    def setup_method(self):
        """Set up test fixtures."""
        self.excel_config = ExcelConfig(
            output_file="test_output.xlsx",
            worksheet_name="TestData",
            auto_fit_columns=True,
            add_timestamp=False  # Disable for testing
        )
        self.writer = ExcelWriter(self.excel_config)
    
    @pytest.mark.asyncio
    async def test_write_new_file(self, sample_scraped_data, temp_dir):
        """Test writing data to a new Excel file."""
        output_file = temp_dir / "new_test.xlsx"
        self.excel_config.output_file = str(output_file)
        
        await self.writer.write_data(sample_scraped_data)
        
        # Verify file was created
        assert output_file.exists()
        
        # Verify content
        df = pd.read_excel(output_file, sheet_name=self.excel_config.worksheet_name)
        assert len(df) == len(sample_scraped_data)
        assert "title" in df.columns
        assert "description" in df.columns
        assert "price" in df.columns
    
    @pytest.mark.asyncio
    async def test_write_empty_data(self, temp_dir):
        """Test writing empty data."""
        output_file = temp_dir / "empty_test.xlsx"
        self.excel_config.output_file = str(output_file)
        
        await self.writer.write_data([])
        
        # Should not create file for empty data
        assert not output_file.exists()
    
    @pytest.mark.asyncio
    async def test_write_with_timestamp(self, sample_scraped_data, temp_dir):
        """Test writing data with timestamp."""
        output_file = temp_dir / "timestamp_test.xlsx"
        self.excel_config.output_file = str(output_file)
        self.excel_config.add_timestamp = True
        
        await self.writer.write_data(sample_scraped_data)
        
        # Verify timestamp column was added
        df = pd.read_excel(output_file, sheet_name=self.excel_config.worksheet_name)
        assert "Timestamp" in df.columns
    
    @pytest.mark.asyncio
    async def test_column_mappings(self, sample_scraped_data, temp_dir):
        """Test column mappings functionality."""
        output_file = temp_dir / "mapped_test.xlsx"
        self.excel_config.output_file = str(output_file)
        self.excel_config.column_mappings = {
            "title": "Product Name",
            "description": "Product Description",
            "price": "Product Price"
        }
        
        await self.writer.write_data(sample_scraped_data)
        
        # Verify mapped column names
        df = pd.read_excel(output_file, sheet_name=self.excel_config.worksheet_name)
        assert "Product Name" in df.columns
        assert "Product Description" in df.columns
        assert "Product Price" in df.columns
        assert "title" not in df.columns  # Original name should be replaced
    
    @pytest.mark.asyncio
    async def test_append_data(self, sample_scraped_data, temp_dir):
        """Test appending data to existing file."""
        output_file = temp_dir / "append_test.xlsx"
        self.excel_config.output_file = str(output_file)
        
        # Write initial data
        await self.writer.write_data(sample_scraped_data[:1])
        
        # Verify initial file
        df_initial = pd.read_excel(output_file, sheet_name=self.excel_config.worksheet_name)
        assert len(df_initial) == 1
        
        # Append more data
        await self.writer.append_data(sample_scraped_data[1:])
        
        # Verify appended data
        df_final = pd.read_excel(output_file, sheet_name=self.excel_config.worksheet_name)
        assert len(df_final) == len(sample_scraped_data)
    
    @pytest.mark.asyncio
    async def test_backup_existing_file(self, sample_scraped_data, temp_dir):
        """Test backup creation for existing files."""
        output_file = temp_dir / "backup_test.xlsx"
        self.excel_config.output_file = str(output_file)
        
        # Create initial file
        initial_data = [{"test": "initial"}]
        df = pd.DataFrame(initial_data)
        df.to_excel(output_file, index=False)
        
        # Write new data (should create backup)
        await self.writer.write_data(sample_scraped_data)
        
        # Check for backup file
        backup_files = list(temp_dir.glob("backup_test_backup_*.xlsx"))
        assert len(backup_files) > 0
    
    def test_excel_formatting_detection(self):
        """Test Excel data formatting detection."""
        # This would test the formatter's ability to detect data types
        # and apply appropriate formatting
        pass
    
    @pytest.mark.asyncio
    async def test_workbook_formatting(self, sample_scraped_data, temp_dir):
        """Test Excel workbook formatting features."""
        output_file = temp_dir / "formatted_test.xlsx"
        self.excel_config.output_file = str(output_file)
        self.excel_config.freeze_header_row = True
        
        await self.writer.write_data(sample_scraped_data)
        
        # Load workbook and check formatting
        workbook = load_workbook(output_file)
        worksheet = workbook[self.excel_config.worksheet_name]
        
        # Check if header row is frozen (freeze_panes should be set)
        assert worksheet.freeze_panes is not None
        
        # Check header formatting
        header_cell = worksheet.cell(row=1, column=1)
        assert header_cell.font.bold is True
        
        workbook.close()
    
    @pytest.mark.asyncio
    async def test_error_handling_invalid_path(self):
        """Test error handling for invalid file paths."""
        # Use a path with invalid characters that will fail on Windows
        self.excel_config.output_file = "C:\\invalid*path?with|bad<chars>.xlsx"
        
        with pytest.raises(Exception):
            await self.writer.write_data([{"test": "data"}])
    
    @pytest.mark.asyncio
    async def test_large_dataset_handling(self, temp_dir):
        """Test handling of large datasets."""
        output_file = temp_dir / "large_test.xlsx"
        self.excel_config.output_file = str(output_file)
        
        # Generate large dataset
        large_data = []
        for i in range(1000):
            large_data.append({
                "id": i,
                "title": f"Item {i}",
                "description": f"Description for item {i}",
                "price": f"${i * 10}.99"
            })
        
        await self.writer.write_data(large_data)
        
        # Verify all data was written
        df = pd.read_excel(output_file, sheet_name=self.excel_config.worksheet_name)
        assert len(df) == 1000
        assert df.iloc[999]["id"] == 999