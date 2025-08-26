"""Excel file writing and management."""

import asyncio
import shutil
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from loguru import logger

from ..core.config import ExcelConfig
from .formatter import ExcelFormatter


class ExcelWriter:
    """Handles Excel file writing with formatting and templates."""
    
    def __init__(self, config: ExcelConfig):
        """Initialize Excel writer.
        
        Args:
            config: Excel configuration
        """
        self.config = config
        self.formatter = ExcelFormatter()
        
    async def write_data(self, data: List[Dict[str, Any]]) -> None:
        """Write scraped data to Excel file.
        
        Args:
            data: List of data items to write
        """
        try:
            logger.info(f"Writing {len(data)} items to Excel file")
            
            # Create output directory if it doesn't exist
            output_path = Path(self.config.output_file)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            # Backup existing file if configured
            if self.config.output_file and output_path.exists():
                await self._backup_existing_file(output_path)
            
            # Convert data to DataFrame
            df = pd.DataFrame(data)
            
            if df.empty:
                logger.warning("No data to write to Excel")
                return
            
            # Add timestamp if configured
            if self.config.add_timestamp:
                df['Timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # Apply column mappings if configured
            if self.config.column_mappings:
                df = df.rename(columns=self.config.column_mappings)
            
            # Create Excel file
            if self.config.template_file and Path(self.config.template_file).exists():
                await self._write_using_template(df)
            else:
                await self._write_new_file(df)
            
            logger.info(f"Successfully wrote data to {self.config.output_file}")
            
        except Exception as e:
            logger.error(f"Error writing Excel file: {e}")
            raise
    
    async def _write_new_file(self, df: pd.DataFrame) -> None:
        """Write data to a new Excel file.
        
        Args:
            df: DataFrame containing data to write
        """
        # Create workbook and worksheet
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = self.config.worksheet_name
        
        # Write data to worksheet
        for row in dataframe_to_rows(df, index=False, header=True):
            worksheet.append(row)
        
        # Apply formatting
        await self._apply_formatting(workbook, worksheet, df)
        
        # Save workbook
        workbook.save(self.config.output_file)
        workbook.close()
    
    async def _write_using_template(self, df: pd.DataFrame) -> None:
        """Write data using an Excel template.
        
        Args:
            df: DataFrame containing data to write
        """
        try:
            # Load template
            workbook = load_workbook(self.config.template_file)
            
            # Get or create worksheet
            if self.config.worksheet_name in workbook.sheetnames:
                worksheet = workbook[self.config.worksheet_name]
                # Clear existing data but keep formatting
                self._clear_worksheet_data(worksheet)
            else:
                worksheet = workbook.create_sheet(self.config.worksheet_name)
            
            # Write data to worksheet
            for row in dataframe_to_rows(df, index=False, header=True):
                worksheet.append(row)
            
            # Apply additional formatting
            await self._apply_formatting(workbook, worksheet, df)
            
            # Save workbook
            workbook.save(self.config.output_file)
            workbook.close()
            
        except Exception as e:
            logger.warning(f"Error using template, falling back to new file: {e}")
            await self._write_new_file(df)
    
    def _clear_worksheet_data(self, worksheet) -> None:
        """Clear data from worksheet while preserving formatting.
        
        Args:
            worksheet: Openpyxl worksheet
        """
        # Find last row and column with data
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        # Clear cell values but keep formatting
        for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.value = None
    
    async def _apply_formatting(self, workbook, worksheet, df: pd.DataFrame) -> None:
        """Apply formatting to the Excel worksheet.
        
        Args:
            workbook: Openpyxl workbook
            worksheet: Openpyxl worksheet
            df: DataFrame with data
        """
        try:
            # Apply header formatting
            await self._format_headers(worksheet, df)
            
            # Apply data formatting
            await self._format_data_cells(worksheet, df)
            
            # Auto-fit columns if configured
            if self.config.auto_fit_columns:
                await self._auto_fit_columns(worksheet, df)
            
            # Freeze header row if configured
            if self.config.freeze_header_row:
                worksheet.freeze_panes = "A2"
            
            # Create table if data exists
            if len(df) > 0:
                await self._create_table(worksheet, df)
            
            logger.debug("Applied Excel formatting successfully")
            
        except Exception as e:
            logger.warning(f"Error applying Excel formatting: {e}")
    
    async def _format_headers(self, worksheet, df: pd.DataFrame) -> None:
        """Format header row.
        
        Args:
            worksheet: Openpyxl worksheet
            df: DataFrame with data
        """
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Apply to header row
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            
            # Add border
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'), 
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border
    
    async def _format_data_cells(self, worksheet, df: pd.DataFrame) -> None:
        """Format data cells based on data type.
        
        Args:
            worksheet: Openpyxl worksheet
            df: DataFrame with data
        """
        # Data alignment
        text_alignment = Alignment(horizontal="left", vertical="center")
        number_alignment = Alignment(horizontal="right", vertical="center")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Apply formatting to data rows
        for row_num in range(2, len(df) + 2):  # Start from row 2 (after header)
            for col_num, (col_name, col_data) in enumerate(df.items(), 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                
                # Apply alignment based on data type
                if col_name.lower() in ['price', 'cost', 'amount', 'total']:
                    cell.alignment = number_alignment
                    # Format as currency if it's a price field
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '$#,##0.00'
                elif col_name.lower() in ['date', 'timestamp', 'created', 'updated']:
                    cell.alignment = center_alignment
                    # Format as date if it's a date field
                    if 'date' in str(cell.value).lower() or 'timestamp' in col_name.lower():
                        cell.number_format = 'yyyy-mm-dd hh:mm:ss'
                else:
                    cell.alignment = text_alignment
                
                # Add border
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'), 
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.border = thin_border
    
    async def _auto_fit_columns(self, worksheet, df: pd.DataFrame) -> None:
        """Auto-fit column widths.
        
        Args:
            worksheet: Openpyxl worksheet
            df: DataFrame with data
        """
        for col_num, column in enumerate(worksheet.columns, 1):
            col_letter = worksheet.cell(row=1, column=col_num).column_letter
            
            # Calculate max width needed
            max_length = 0
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            # Set column width (with some padding)
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[col_letter].width = adjusted_width
    
    async def _create_table(self, worksheet, df: pd.DataFrame) -> None:
        """Create an Excel table for better data presentation.
        
        Args:
            worksheet: Openpyxl worksheet
            df: DataFrame with data
        """
        try:
            # Define table range
            table_range = f"A1:{worksheet.cell(row=len(df) + 1, column=len(df.columns)).coordinate}"
            
            # Create table
            table = Table(displayName="ScrapedData", ref=table_range)
            
            # Add table style
            style = TableStyleInfo(
                name="TableStyleMedium9", 
                showFirstColumn=False,
                showLastColumn=False, 
                showRowStripes=True, 
                showColumnStripes=True
            )
            table.tableStyleInfo = style
            
            # Add table to worksheet
            worksheet.add_table(table)
            
        except Exception as e:
            logger.warning(f"Error creating Excel table: {e}")
    
    async def _backup_existing_file(self, file_path: Path) -> None:
        """Create backup of existing Excel file.
        
        Args:
            file_path: Path to existing file
        """
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_path = file_path.with_name(f"{file_path.stem}_backup_{timestamp}{file_path.suffix}")
            
            shutil.copy2(file_path, backup_path)
            logger.info(f"Created backup: {backup_path}")
            
        except Exception as e:
            logger.warning(f"Error creating backup: {e}")
    
    async def append_data(self, data: List[Dict[str, Any]]) -> None:
        """Append data to existing Excel file.
        
        Args:
            data: List of data items to append
        """
        try:
            output_path = Path(self.config.output_file)
            
            if output_path.exists():
                # Load existing workbook
                workbook = load_workbook(self.config.output_file)
                worksheet = workbook[self.config.worksheet_name]
                
                # Find next empty row
                next_row = worksheet.max_row + 1
                
                # Convert data to DataFrame
                df = pd.DataFrame(data)
                
                # Add timestamp if configured
                if self.config.add_timestamp:
                    df['Timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                
                # Append data
                for row in dataframe_to_rows(df, index=False, header=False):
                    worksheet.append(row)
                
                # Apply formatting to new rows
                await self._format_new_rows(worksheet, df, next_row)
                
                # Save workbook
                workbook.save(self.config.output_file)
                workbook.close()
                
                logger.info(f"Appended {len(data)} rows to existing file")
                
            else:
                # File doesn't exist, create new one
                await self.write_data(data)
                
        except Exception as e:
            logger.error(f"Error appending data to Excel file: {e}")
            raise
    
    async def _format_new_rows(self, worksheet, df: pd.DataFrame, start_row: int) -> None:
        """Apply formatting to newly added rows.
        
        Args:
            worksheet: Openpyxl worksheet
            df: DataFrame with new data
            start_row: Starting row number for new data
        """
        # Apply same formatting as regular data cells
        text_alignment = Alignment(horizontal="left", vertical="center")
        number_alignment = Alignment(horizontal="right", vertical="center")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        for row_offset, (_, row_data) in enumerate(df.iterrows()):
            row_num = start_row + row_offset
            
            for col_num, (col_name, col_value) in enumerate(row_data.items(), 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                
                # Apply alignment based on data type
                if col_name.lower() in ['price', 'cost', 'amount', 'total']:
                    cell.alignment = number_alignment
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '$#,##0.00'
                elif col_name.lower() in ['date', 'timestamp', 'created', 'updated']:
                    cell.alignment = center_alignment
                else:
                    cell.alignment = text_alignment
                
                # Add border
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'), 
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.border = thin_border